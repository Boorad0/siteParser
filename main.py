from selenium import webdriver
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = "Ссылка на продукт"
ws["B1"] ="Название модели"
ws["C1"] ="Х/р"
ws["D1"] = "Цена"
ws["E1"] ="Ссылка на картинку"
ws["F1"] ="Описание"
def get_html():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://iotvega.com/product")
    page = driver.page_source
    soup = BeautifulSoup(page, "html.parser")
    all_href = soup.find_all("a", class_="main-container")
    number = 2
    for what in all_href:
        ws[f"A{number}"] = "https://iotvega.com/"+what.get("href")
        ws[f"B{number}"] =what.find_all("h2")[0].text
        ws[f"C{number}"] =what.find_all("li")[0].text
        ws[f"D{number}"] = what.find_all("span", {"class":"price_item"})[0].text.replace("  ", "")
        ws[f"E{number}"] ="https://iotvega.com/"+what.find("img").attrs["src"][2:]
        number+=1
    driver.close()
    driver.quit()
def get_inforamation():
    kolumn = 2
    driver = webdriver.Chrome()
    driver.maximize_window()
    while ws[f"A{kolumn}"].value !=None:
        driver.get(ws[f"A{kolumn}"].value)
        page = driver.page_source
        soup = BeautifulSoup(page, "html.parser")
        inf = soup.find("section", {"class":"padding-top"}).find("div",class_="col-sm-12").text.replace("  ", "")
        ws[f"F{kolumn}"] = inf
        kolumn +=1
def main():

    get_html()
    get_inforamation()
    wb.save('test.xlsx')

if __name__ == "__main__":
    main()